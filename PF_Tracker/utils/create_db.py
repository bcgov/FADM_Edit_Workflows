######################################################################
## Purpose: Create SQLite database from Admin Boundaries Provincial Forest spreadsheet
## Date: April, 2021
## Author: Brett Edwards
###############################################################################
import openpyxl
import sqlite3
import datetime


CREATE_TABLE = """ CREATE TABLE IF NOT EXISTS pf_tracking (
                    id INTEGER PRIMARY KEY,
                    prov_forest_name TEXT,
                    lands_file TEXT,
                    prov_forest_tracking_num INTEGER,
                    ministerial_order TEXT,
                    date_received TEXT,
                    legal_description TEXT,
                    area_ha REAL,
                    sup TEXT,
                    approval_in_principle TEXT,
                    deletion_approved_date TEXT,
                    ha_deleted REAL,
                    purpose TEXT,
                    date_to_mfr TEXT,
                    date_signed TEXT,
                    district TEXT,
                    date_last_edited TEXT,
                    last_edited_by TEXT,
                    notes TEXT,
                    FOREIGN KEY(prov_forest_name) REFERENCES pf_region_file(prov_forest_name)
);"""

CREATE_REGION_FILE_TABLE = """ CREATE TABLE IF NOT EXISTS pf_region_file (
    prov_forest_name TEXT PRIMARY KEY NOT NULL,
    region TEXT NOT NULL,
    file_number TEXT
);
"""

INSERT_PF_ROW = """
    INSERT INTO pf_tracking (prov_forest_name, lands_file, prov_forest_tracking_num, ministerial_order, date_received, legal_description, area_ha, sup, approval_in_principle, deletion_approved_date, ha_deleted, purpose, date_to_mfr, date_signed, district, notes) 
    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
"""

SELECT_DISTINCT_PFS = "SELECT DISTINCT prov_forest_name FROM pf_tracking;"

INSERT_REGION_FILE = "INSERT INTO pf_region_file (prov_forest_name, region, file_number) VALUES (?, ?, ?)"


def insert_data(pf_tuple):
    with connection:
        connection.execute(INSERT_PF_ROW, pf_tuple)


def create_tables():
    with connection:
        connection.execute(CREATE_TABLE)
        connection.execute(CREATE_REGION_FILE_TABLE)


def insert_region_data(pf_name, region, file_number):
    with connection:
        connection.execute(INSERT_REGION_FILE, (pf_name, region, file_number))


def get_pfs():
    with connection:
        cursor = connection.cursor()
        cursor.execute(SELECT_DISTINCT_PFS)
        return cursor.fetchall()

date = datetime.datetime.now().strftime("%Y-%m-%d")
connection = sqlite3.connect(Rf"C:\Projects\PF_db\Provincial_Forest_Tracking_{date}.db")

###########################################################################################################################################################


workbook = R'Q:\projects\Mflnro\FADM_Provincial_Forests\_Procedures\Current Record of Changes to PF by MO.xlsx'

book = openpyxl.load_workbook(workbook, data_only=True)
sheets = book.sheetnames

create_tables()



####################################################################################################
# insert forest region and file number information into db table 'pf_region_file'

for sheet in sheets:
    ws = book[sheet]

    pf_name = sheet

    forest_region = ws["A6"].value  
    file_number = ws["A7"].value

    _, forest_region = forest_region.split(': ')
    _, file_number = file_number.split(': ')

    insert_region_data(pf_name, forest_region, file_number)




##########################################################################################
# insert all data rows from spreadsheet into db

for sheet in sheets:
    ws = book[sheet]

    # start row
    row = 10
    row_empty = False

    while not row_empty:
        curr_row = f"A{row}:N{row}"
        data = ws[curr_row]

        row_vals = [cell.value for cell in data[0]]

        # openpyxl returns datetime objects from dates in excel. We want a string to insert into the db (sqlite doesn't have a 'DATE' type)
        for idx, val in enumerate(row_vals):
            if isinstance(val, datetime.datetime):
                row_vals[idx] = val.strftime('%Y-%m-%d')

        # PFT and MO are in a single entry in the spreadsheet, but we actually want them split up
        pft_mo = row_vals[1]
        if type(pft_mo) == str and '(' in pft_mo:
            pft, mo = pft_mo.split('(')
            mo = mo.strip("()")
            row_vals[1] = int(pft)
            row_vals.insert(2, mo)
        else:
            row_vals.insert(2, None)

        # check if the row is empty, increment the row # until it is        
        row_empty = all(cell is None for cell in row_vals)
        row += 1

        
        if not row_empty:
            row_vals.insert(0, sheet)
            # print(len(row_vals))
            # print(row_vals)

            insert_data(row_vals)














